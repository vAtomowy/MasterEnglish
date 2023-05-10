# Master english - program to learn ENG language 
# Artur Bereit / 8.01.23 & 10.05.23

# Qt5 
from PyQt5.QtWidgets import QApplication, QAction, QSystemTrayIcon, QMenu
from PyQt5.QtGui import QIcon

# System 
import sys
import os

# Keyboard command
import keyboard 

# Other math lib 
import math 
import random

# Excel document 
import openpyxl

# copy clip?
import pyautogui as pya
import pyperclip  # `handy` cross-platform clipboard text handler
import time

#notifications
from win10toast import ToastNotifier 

#translator 
from deep_translator import GoogleTranslator

class word: 
    last_string = " "
    input_string =""
    to_translate =""
    translated =""

#Current Directory 
CURRENT_DIRECTORY = os.path.dirname(os.path.realpath(__file__))

def copy_clipboard():
    pya.hotkey('ctrl', 'c')
    time.sleep(.05)  # ctrl-c is usually very fast but your program may execute faster
    return pyperclip.paste()

def event(): 
    # print("| ` + 1 | -> detected")
    word.input_string = copy_clipboard()
    if(word.last_string == word.input_string):
        return 
    else:
        word.last_string = word.input_string
        # Start by opening the spreadsheet and selecting the main sheet
        filename = (os.path.join(CURRENT_DIRECTORY, "data.xlsx"))
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active

        # print(sheet.max_row)
        last_num = sheet.max_row; 
        sheet.cell(row = last_num+1, column = 1, value = word.input_string)

        word.to_translate = ""; 
        word.translated = "";
        try:
            word.to_translate = word.input_string;
            word.translated = GoogleTranslator(source='en', target='pl').translate(word.to_translate);
            sheet.cell(row = last_num+1, column = 2, value = word.translated)
            toaster = ToastNotifier()
            toaster.show_toast("Master english App","Added a word to learn", duration=0.7)  
        except: 
            toaster = ToastNotifier()
            toaster.show_toast("Master english App","Uncorrect word !", duration=0.7)      
        
        # Save the spreadsheet
        workbook.save(filename=filename)

        return

def quit_event():
    print("wykryciee")
    os._exit(0);

# keyboard.add_hotkey('ctrl+`', event)
keyboard.add_hotkey('`+1', event);

#end app
keyboard.add_hotkey('`+0', quit_event);

app = QApplication([])
app.setQuitOnLastWindowClosed(False)
  
# Adding an icon
icon = QIcon(os.path.join(CURRENT_DIRECTORY, "g.png"))
  
# Adding item on the menu bar
tray = QSystemTrayIcon()
tray.setIcon(icon)
tray.setVisible(True)
  
# Creating the options
menu = QMenu()
option1 = QAction("Word List")
option2 = QAction("About App")
menu.addAction(option1)
menu.addAction(option2)
  
# To quit the app
quit = QAction("Quit")
quit.triggered.connect(app.quit)
menu.addAction(quit)
  
# Adding options to the System Tray
tray.setContextMenu(menu)
  
app.exec_()
